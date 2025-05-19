"""
Excel解析器，负责解析Excel电子表格并提取文本内容和元数据
"""
import os
from typing import Dict, Any, List, Optional
import openpyxl
from core.utils.logger import get_logger

logger = get_logger(__name__)


class XlsxParser:
    """Excel电子表格解析器，提取xlsx文件的文本内容和元数据"""
    
    def __init__(self):
        """初始化Excel解析器"""
        pass
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        解析Excel电子表格
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 包含标题、内容和元数据的字典
        """
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return {}
        
        try:
            result = {
                'title': os.path.basename(file_path),
                'content': '',
                'metadata': {}
            }
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 提取元数据
            metadata = {}
            metadata['sheet_names'] = workbook.sheetnames
            metadata['sheet_count'] = len(workbook.sheetnames)
            
            if workbook.properties:
                if workbook.properties.title:
                    metadata['title'] = workbook.properties.title
                    result['title'] = workbook.properties.title
                if workbook.properties.creator:
                    metadata['creator'] = workbook.properties.creator
                if workbook.properties.created:
                    metadata['created'] = str(workbook.properties.created)
                if workbook.properties.modified:
                    metadata['modified'] = str(workbook.properties.modified)
                if workbook.properties.subject:
                    metadata['subject'] = workbook.properties.subject
                if workbook.properties.keywords:
                    metadata['keywords'] = workbook.properties.keywords
                if workbook.properties.category:
                    metadata['category'] = workbook.properties.category
            
            result['metadata'] = metadata
            
            # 提取表格内容
            content_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content_parts.append(f"工作表: {sheet_name}")
                
                # 计算数据行和列的范围
                min_row, max_row = 1, sheet.max_row
                min_col, max_col = 1, sheet.max_column
                
                # 防止处理过大的表格导致性能问题
                if max_row > 1000:
                    logger.warning(f"表格 {sheet_name} 行数过多 ({max_row}), 仅处理前1000行")
                    max_row = 1000
                
                if max_col > 100:
                    logger.warning(f"表格 {sheet_name} 列数过多 ({max_col}), 仅处理前100列")
                    max_col = 100
                
                # 提取表格数据
                rows_data = []
                for row in range(min_row, max_row + 1):
                    row_values = []
                    for col in range(min_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        value = cell.value
                        row_values.append(str(value) if value is not None else "")
                    
                    # 跳过完全为空的行
                    if any(v.strip() for v in row_values):
                        rows_data.append(" | ".join(row_values))
                
                content_parts.append("\n".join(rows_data))
            
            result['content'] = "\n\n".join(content_parts)
            
            return result
            
        except Exception as e:
            logger.error(f"解析Excel文件失败: {file_path}, 错误: {str(e)}")
            return {}
    
    def extract_charts(self, file_path: str) -> Dict[str, Any]:
        """
        从Excel提取图表数据（可选功能）
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 图表数据字典
        """
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return {}
        
        try:
            chart_data = {}
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 注意：openpyxl的图表提取功能有限
                # 这里只是一个简化的实现，可能需要扩展
                if sheet._charts:
                    chart_list = []
                    for idx, chart in enumerate(sheet._charts):
                        chart_info = {
                            'type': type(chart).__name__,
                            'title': getattr(chart.title, 'text', f'Chart {idx+1}')
                        }
                        chart_list.append(chart_info)
                    
                    if chart_list:
                        chart_data[sheet_name] = chart_list
            
            return chart_data
            
        except Exception as e:
            logger.error(f"从Excel提取图表失败: {file_path}, 错误: {str(e)}")
            return {}